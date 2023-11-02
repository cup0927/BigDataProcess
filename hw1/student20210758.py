#!/usr/bin/env python
# coding: utf-8

# In[8]:


#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active 

#현재 데이터의 행과 열의 수 파악하기
max_rows = ws.max_row
max_cols = ws.max_column

#데이터 읽어서 total 계산하기
for i in range (2, max_rows+1):
    total =0
    for j in range(3, 8):
        if(j==3):
            total+= int(ws.cell(row = i, column = j).value) * 0.3
        elif(j==4):
            total+= int(ws.cell(row = i, column = j).value) * 0.35
        elif(j==5):
            total+= int(ws.cell(row = i, column = j).value) * 0.34
        elif(j==6):
            total+= int(ws.cell(row = i, column = j).value)
        else:
            ws.cell(row = i, column = j, value= total)  

#데이터 읽어서 grade 계산하기
d1 = dict()

numberOfA = int((max_rows - 1) * 0.3)
numberOfAPlus = numberOfA // 2
numberOfAZero = numberOfA - numberOfAPlus
numberOfB = int((max_rows - 1) * 0.7) - numberOfA
numberOfBPlus = numberOfB // 2
numberOfBZero = numberOfB - numberOfBPlus
numberOfC = int((max_rows -1) - int((max_rows - 1) * 0.7))
numberOfCPlus = numberOfC // 2
numberOfCZero = numberOfC - numberOfCPlus

countAZero, countAPlus, countBZero, countBPlus, countCZero, countCPlus = 0, 0, 0, 0, 0, 0
# row index 와 total 딕셔너리에 저장 및 내림차순 정렬
for i in range (2, max_rows+1):
    d1[i] = ws.cell(row = i, column = 7).value
    
d1 = sorted(d1.items(), key=lambda x: x[1], reverse=True)
total_dic = dict(map(reversed, d1))

#디버깅용
# print(numberOfAPlus)
# print(numberOfAZero)
# print(numberOfB)
# print(numberOfBPlus)
# print(numberOfBZero)
# print(numberOfCPlus)
# print(numberOfCZero)

for i in total_dic.values():
    if int(ws.cell(row = i, column = 7).value) < 40:
        ws.cell(row = i, column = 8, value= 'F')
        continue     
    elif countAPlus < numberOfAPlus :
        ws.cell(row = i, column = 8, value= 'A+')
        countAPlus+=1
    elif countAZero < numberOfAZero :
        ws.cell(row = i, column = 8, value= 'A') 
        countAZero+=1
    elif countBPlus < numberOfBPlus :
        ws.cell(row = i, column = 8, value= 'B+')
        countBPlus+=1
    elif countBZero < numberOfBZero :
        ws.cell(row = i, column = 8, value= 'B') 
        countBZero+=1
    elif countCPlus < numberOfCPlus :
        ws.cell(row = i, column = 8, value= 'C+')
        countCPlus+=1
    elif countCZero < numberOfCZero :
        ws.cell(row = i, column = 8, value= 'C') 
        countCZero+=1
    
wb.save('student.xlsx')

